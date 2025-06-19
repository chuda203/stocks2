import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import logging

def create_excel_download(stock_symbol, fitting_dates, fitting_prices, Fitting_S_n_list, 
                        forecast_dates, S_forecast, actual_forecast_prices):
    try:
        # Create fitting dataframe
        fitting_df = pd.DataFrame({
            'Date': fitting_dates,
            'Actual_Price': fitting_prices,
            'Fitted_Price': Fitting_S_n_list[:len(fitting_dates)],
            'Forecast_Price': None,
            'Type': 'Fitting'
        })
        
        # Create forecast dataframe if available
        if forecast_dates and S_forecast and actual_forecast_prices:
            min_len = min(len(forecast_dates), len(S_forecast), len(actual_forecast_prices))
            forecast_df = pd.DataFrame({
                'Date': forecast_dates[:min_len],
                'Actual_Price': actual_forecast_prices[:min_len],
                'Fitted_Price': None,
                'Forecast_Price': S_forecast[:min_len],
                'Type': 'Forecast'
            })
            combined_df = pd.concat([fitting_df, forecast_df], ignore_index=True)
        else:
            combined_df = fitting_df
        
        # Create Excel file
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{stock_symbol}_Analysis"
        
        # Add headers
        ws['A1'] = f"Stock Analysis Report - {stock_symbol}"
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Fitting Period: {fitting_dates[0].strftime('%Y-%m-%d')} to {fitting_dates[-1].strftime('%Y-%m-%d')}"
        if forecast_dates:
            ws['A4'] = f"Forecast Period: {forecast_dates[0].strftime('%Y-%m-%d')} to {forecast_dates[-1].strftime('%Y-%m-%d')}"
        ws['A5'] = ""
        
        title_font = Font(size=14, bold=True)
        ws['A1'].font = title_font
        
        # Add data headers
        headers = ['Date', 'Actual Price', 'Fitted Price', 'Forecast Price', 'Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_idx, (_, row) in enumerate(combined_df.iterrows(), 7):
            if pd.notna(row['Date']):
                if hasattr(row['Date'], 'strftime'):
                    ws.cell(row=row_idx, column=1, value=row['Date'].strftime('%Y-%m-%d'))
                else:
                    ws.cell(row=row_idx, column=1, value=str(row['Date']))
            else:
                ws.cell(row=row_idx, column=1, value="")
                
            ws.cell(row=row_idx, column=2, value=row['Actual_Price'])
            ws.cell(row=row_idx, column=3, value=row['Fitted_Price'])
            ws.cell(row=row_idx, column=4, value=row['Forecast_Price'])
            ws.cell(row=row_idx, column=5, value=row['Type'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        logging.error(f"Error in create_excel_download: {e}")
        raise e